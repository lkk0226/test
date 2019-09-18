# coding:utf-8

import openpyxl
from common.logFormatter import get_logger

logger = get_logger("F:\\pythonPjo\\testExcelCase\\log\\readExcel.txt")

class ReadExcel:

    def readExcel(fileName,SheetName):
        # 创建对象
        wb = openpyxl.load_workbook(fileName)
        # 创建sheet对象
        ws = wb[SheetName]
        rows = ws.max_row
        cols = ws.max_column
        test_case = []

        if rows > 1:
            for row in range(1, rows + 1):
                sub_date = {}
                sub_date['rownum'] = row
                sub_date['casename'] = ws.cell(row, 1).value
                sub_date['method'] = ws.cell(row, 2).value
                sub_date['url'] = ws.cell(row, 3).value
                if isinstance(ws.cell(row, 4).value, str):
                    sub_date['hearder'] = ws.cell(row, 4).value
                else:
                    sub_date['hearder']=None
                if isinstance(ws.cell(row, 5).value, str):
                    sub_date['param'] = ws.cell(row, 5).value
                else:
                    sub_date['param'] = None
                sub_date['expectRes'] = ws.cell(row, 6).value
                sub_date['flag'] = ws.cell(row, 7).value

                if sub_date['flag'] == 1:
                    test_case.append(sub_date)
                else:
                    continue

            logger.info("从excel读取到的测试用例list:%s" % test_case)
            return test_case
        else:
            logger.warn("表格为空")
            return None


if __name__ == '__main__':
    s=ReadExcel.readExcel("F:\\pythonPjo\\testExcelCase\\data\\login.xlsx","test")
    logger.info("运行结果：%s" % s)

# testExcel().readExcel()






